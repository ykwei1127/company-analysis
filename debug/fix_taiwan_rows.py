"""
針對 Wistron / Delta Electronics / NVIDIA / TSMC 的 Taiwan 筆，
用 _IN240 URL 重新爬取，並更新 output/ 下最新的 xlsx 和 csv。
"""
import sys, os, glob, re, time
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from glassdoor_scraper_unified import GlassdoorScraper
import pandas as pd

PORT = 9222

# 公司 ID 對應表（從 matched json 推斷）
# URL 格式：EI_IE{id}.0,{name_len}_IL.{il_start},{il_end}_IN240.htm
# 直接從現有 Taipei URL 替換 IC code 為 IN240

COMPANIES = [
    {
        'name': 'Wistron',
        'taipei_url': 'https://www.glassdoor.com/Reviews/Wistron-Taipei-Reviews-EI_IE15218.0,7_IL.8,14_IC3271041.htm',
    },
    {
        'name': 'Delta Electronics',
        'taipei_url': 'https://www.glassdoor.com/Reviews/Delta-Electronics-Taipei-Reviews-EI_IE41146.0,17_IL.18,24_IC3271041.htm',
    },
    {
        'name': 'NVIDIA',
        'taipei_url': 'https://www.glassdoor.com/Reviews/NVIDIA-Hsinchu-Reviews-EI_IE7633.0,6_IL.7,14_IC4915323.htm',
    },
    {
        'name': 'TSMC',
        'taipei_url': 'https://www.glassdoor.com/Reviews/TSMC-Hsinchu-Reviews-EI_IE4130.0,4_IL.5,12_IC4915323.htm',
    },
]

def build_taiwan_url(taipei_url):
    """把城市 IC URL 改成 Taiwan IN240 URL"""
    # 替換 CityName-Reviews 為 Taiwan-Reviews
    url = re.sub(r'/Reviews/(\w[^/]*?)-(?:Taipei|Hsinchu)-Reviews-', '/Reviews/\\1-Taiwan-Reviews-', taipei_url)
    # 替換 IL 後面的 _IC... 為 _IN240
    url = re.sub(r'(_IL\.\d+,\d+)_IC\d+\.htm', r'\1_IN240.htm', url)
    return url

print("=== Taiwan IN240 URL 預覽 ===")
for c in COMPANIES:
    tw_url = build_taiwan_url(c['taipei_url'])
    c['taiwan_url'] = tw_url
    print(f"{c['name']}: {tw_url}")

print("\n開始爬取...\n")
scraper = GlassdoorScraper(mode='manual', chrome_debugger_port=PORT)

results = {}
for c in COMPANIES:
    display = f"{c['name']} - Taiwan"
    data = scraper.extract_rating_data(c['taiwan_url'], display)
    results[c['name']] = data
    time.sleep(3)
    print()

scraper.close()

# 找最新的 xlsx
xlsx_files = sorted(glob.glob('output/glassdoor_ratings*.xlsx'))
if not xlsx_files:
    print("找不到 output/*.xlsx")
    sys.exit(1)
latest_xlsx = xlsx_files[-1]
latest_csv = os.path.splitext(latest_xlsx)[0] + '.csv'
print(f"\n更新: {latest_xlsx}")

df = pd.read_excel(latest_xlsx)

for c in COMPANIES:
    name = c['name']
    data = results.get(name)
    if not data:
        print(f"  ✗ {name} 爬取失敗，跳過")
        continue

    # 找到 baseline_location == Taiwan 且 Company 開頭符合的那行
    mask = df['Company'].str.startswith(name) & (df['Baseline Location'] == 'Taiwan')
    if mask.sum() == 0:
        print(f"  ✗ 找不到 {name} - Taiwan 列")
        continue

    idx = df[mask].index[0]
    for col in ['Overall', 'Recommend', 'CEO Approval', 'Total Reviews',
                'Diversity & Inclusion', 'Work/Life Balance',
                'Compensation and Benefits', 'Culture & Values',
                'Career Opportunities', 'Senior Management']:
        if col in data and data[col] is not None:
            df.at[idx, col] = data[col]

    df.at[idx, 'Actual City'] = 'Taiwan'
    df.at[idx, 'Review URL'] = c['taiwan_url']
    print(f"  ✓ {name} - Taiwan 更新完成  Overall={data.get('Overall')}")

# 存回 xlsx + csv
df.to_excel(latest_xlsx, index=False, engine='openpyxl')
df.to_csv(latest_csv, index=False, encoding='utf-8-sig')

# 補套格式
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

wb = load_workbook(latest_xlsx)
ws = wb.active

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

col_widths = {'A': 30, 'B': 22, 'C': 22, 'D': 10, 'E': 12,
              'F': 14, 'G': 20, 'H': 18, 'I': 24, 'J': 16,
              'K': 22, 'L': 18}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save(latest_xlsx)
print(f"\n✓ 已存回 {latest_xlsx}")
print(f"✓ 已存回 {latest_csv}")
