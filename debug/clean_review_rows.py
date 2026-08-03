"""
Clean a Glassdoor ratings workbook by validating each Review URL with the same
country/city page checks used in company_finder.py.

Default input:
  output/glassdoor_ratings_20260624_2242.xlsx
Default output:
  output/glassdoor_ratings_20260624_2242_cleaned.xlsx

Default Chrome debug ports:
  9225, 9226, 9227
"""
from __future__ import annotations

import argparse
import queue
import re
import sys
import threading
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from company_finder import CompanyFinder  # noqa: E402

DEFAULT_INPUT = PROJECT_ROOT / "output" / "glassdoor_ratings_20260624_2242.xlsx"
DEFAULT_PORTS = [9225, 9226, 9227]
BLOCK_RETRY_SECONDS = 60
PAGE_WAIT_SECONDS = 8
ROW_DELAY_SECONDS = 1.0

BLOCK_TITLES = {"just a moment...", "attention required"}
BLOCK_INDICATORS = [
    "humans only",
    "checking your browser",
    "cf-challenge-running",
    "verify you are human",
    "cf-turnstile",
    "challenge-platform",
]


class StopRequested(Exception):
    pass


def _sleep_interruptible(stop_event: threading.Event, seconds: float, step: float = 0.5):
    end = time.time() + seconds
    while time.time() < end:
        if stop_event.is_set():
            raise StopRequested()
        time.sleep(min(step, max(0.0, end - time.time())))


def _wait_ready(driver, stop_event: threading.Event, timeout=PAGE_WAIT_SECONDS):
    end = time.time() + timeout
    while time.time() < end:
        if stop_event.is_set():
            raise StopRequested()
        try:
            if driver.execute_script("return document.readyState") == "complete":
                return
        except Exception:
            pass
        time.sleep(0.5)


def _clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none", "null"} else text


def _company_base(value) -> str:
    return _clean_text(value).split(" - ")[0].strip()


def _has_numeric_value(value) -> bool:
    text = _clean_text(value)
    if not text:
        return False
    return not pd.isna(pd.to_numeric(text, errors="coerce"))


def _normalize_slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def _extract_review_label(url: str, company_name: str) -> str:
    """Infer the location label embedded in the review URL."""
    if not url:
        return ""
    match = re.search(r"/Reviews/(.+?)-Reviews-", url)
    if not match:
        return ""
    path = match.group(1).strip()
    if not path:
        return ""

    company_slug = _normalize_slug(company_name)
    if company_slug and path.lower() == company_slug:
        return "Global"
    if company_slug and path.lower().startswith(company_slug + "-"):
        label = path[len(company_slug) + 1 :]
    else:
        label = path

    label = label.replace("-", " ").strip()
    return label or "Global"


def _row_source_mode(row: pd.Series) -> str:
    return _clean_text(row.get("Source Mode")).lower()


def _location_core_label(value: str) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    for sep in [",", " - ", "–", "—"]:
        if sep in text:
            text = text.split(sep)[0].strip()
            break
    return text


def _infer_expected_label(row: pd.Series) -> str:
    url = _clean_text(row.get("Review URL"))
    source_mode = _row_source_mode(row)

    country_level = (
        "_IN" in url
        or source_mode == "country"
        or "country_fb" in source_mode
        or source_mode.startswith("country+")
        or source_mode.startswith("city+country_fb")
        or source_mode.startswith("office+country_fb")
    )

    columns = ["Country", "Baseline Location", "Actual City"] if country_level else ["Baseline Location", "Actual City", "Country"]

    for col in columns:
        value = _clean_text(row.get(col))
        if not value:
            continue
        if country_level:
            return value
        core = _location_core_label(value)
        if core:
            return core

    company = _company_base(row.get("Company"))
    return _extract_review_label(url, company)


def _infer_expected_company(row: pd.Series) -> str:
    return _company_base(row.get("Company"))


def _extract_heading(driver) -> str:
    for sel in ["h1", "h2", '[class*="heading"]', '[class*="Heading"]']:
        try:
            elems = driver.find_elements(By.CSS_SELECTOR, sel)
            for elem in elems:
                text = elem.text.strip()
                if text and 3 < len(text) < 80:
                    return text
        except Exception:
            continue
    return ""


def _is_blocked(driver) -> bool:
    try:
        title = driver.title.strip().lower()
        if title in BLOCK_TITLES or "security" in title:
            return True
        page_source = (driver.page_source or "")[:2000].lower()
        if "help us protect glassdoor" in page_source:
            return True
        return any(ind in page_source for ind in BLOCK_INDICATORS)
    except Exception:
        return False


def _validate_global_page(driver, url: str) -> tuple[bool, str]:
    final_url = driver.current_url.split("?")[0]
    eid_match = re.search(r"E(?:I_IE)?(\d+)", url)
    expected_id = eid_match.group(1) if eid_match else ""
    if expected_id and f"E{expected_id}" not in final_url:
        return False, f"global id mismatch: expected E{expected_id}, got {final_url}"
    if "Location" in final_url:
        return False, f"redirected to location page: {final_url}"
    return True, ""


def _validate_row(finder: CompanyFinder, row: pd.Series, stop_event: threading.Event) -> tuple[bool, str]:
    if stop_event.is_set():
        raise StopRequested()
    if not _has_numeric_value(row.get("Overall")):
        return False, "missing Overall"

    url = _clean_text(row.get("Review URL"))
    if not url:
        return False, "missing Review URL"

    driver = finder.driver
    driver.get(url)
    _wait_ready(driver, stop_event)

    if stop_event.is_set():
        raise StopRequested()

    if _is_blocked(driver):
        _sleep_interruptible(stop_event, BLOCK_RETRY_SECONDS)
        driver.get(url)
        _wait_ready(driver, stop_event)
        if _is_blocked(driver):
            return False, "blocked by Cloudflare / rate limit"

    if stop_event.is_set():
        raise StopRequested()

    expected_label = _infer_expected_label(row)
    expected_company = _infer_expected_company(row)
    page_title = driver.title.strip()
    page_heading = _extract_heading(driver)
    final_url = driver.current_url.split("?")[0]

    if expected_label and expected_label.lower() == "global":
        ok, reason = _validate_global_page(driver, url)
        if not ok:
            return False, reason
        return True, ""

    ok = finder._validate_review_page(  # noqa: SLF001
        expected_label,
        page_title,
        page_heading,
        final_url,
        context="Excel clean",
        expected_company=expected_company,
        allow_company_heading=True,
    )
    if not ok:
        return False, (
            f"mismatch: expected={expected_label!r}, title={page_title!r}, "
            f"heading={page_heading!r}, url={final_url!r}"
        )
    return True, ""


def _worker(port: int, task_queue: queue.Queue, results: dict, lock: threading.Lock, errors: list[str], stop_event: threading.Event, total_rows: int, progress: dict):
    finder = None
    try:
        finder = CompanyFinder(port)
        try:
            finder.driver.set_page_load_timeout(20)
        except Exception:
            pass
        while not stop_event.is_set():
            try:
                idx, row = task_queue.get_nowait()
            except queue.Empty:
                break

            try:
                keep, reason = _validate_row(finder, row, stop_event)
            except StopRequested:
                task_queue.task_done()
                break
            except Exception as exc:
                keep, reason = False, f"exception: {exc}"

            with lock:
                results[idx] = {
                    "keep": keep,
                    "reason": reason,
                    "port": port,
                    "company": _clean_text(row.get("Company")),
                    "baseline": _clean_text(row.get("Baseline Location")),
                    "overall": _clean_text(row.get("Overall")),
                    "url": _clean_text(row.get("Review URL")),
                }
                progress["done"] = progress.get("done", 0) + 1
                print(
                    f"[{progress['done']}/{total_rows}] row {idx + 2}: "
                    f"{_clean_text(row.get('Company'))} | {_clean_text(row.get('Baseline Location'))}",
                    flush=True,
                )
            task_queue.task_done()
            if not stop_event.is_set():
                _sleep_interruptible(stop_event, ROW_DELAY_SECONDS)
    except Exception as exc:
        with lock:
            errors.append(f"port {port}: {exc}")
    finally:
        try:
            if finder:
                finder.driver.quit()
        except Exception:
            pass


def _style_output_xlsx(path: Path):
    wb = load_workbook(path)
    ws = wb.active

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    review_url_col = None
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if str(cell.value).strip() == "Review URL":
            review_url_col = cell.column_letter

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            text = "" if cell.value is None else str(cell.value)
            letter = cell.column_letter
            widths[letter] = max(widths.get(letter, 0), len(text))
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for letter, max_len in widths.items():
        if review_url_col and letter == review_url_col:
            ws.column_dimensions[letter].width = 60
        else:
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)

    wb.save(path)


def clean_workbook(input_path: Path, output_path: Path, ports: list[int]) -> dict:
    if input_path.suffix.lower() == ".xlsx":
        source_wb = load_workbook(input_path, read_only=True)
        sheet_name = source_wb.sheetnames[0]
        source_wb.close()
        df = pd.read_excel(input_path)
    elif input_path.suffix.lower() == ".csv":
        sheet_name = "Sheet1"
        df = pd.read_csv(input_path)
    else:
        raise ValueError(f"Unsupported input file type: {input_path.suffix}")

    if "Review URL" not in df.columns:
        raise ValueError("Missing required column: Review URL")
    if "Overall" not in df.columns:
        raise ValueError("Missing required column: Overall")

    task_queue: queue.Queue = queue.Queue()
    for idx, row in df.iterrows():
        task_queue.put((idx, row))

    total_rows = len(df)
    print(f"Total rows: {total_rows}")

    results: dict = {}
    lock = threading.Lock()
    errors: list[str] = []
    stop_event = threading.Event()
    progress = {"done": 0}

    threads = []
    for port in ports:
        t = threading.Thread(
            target=_worker,
            args=(port, task_queue, results, lock, errors, stop_event, total_rows, progress),
            daemon=True,
        )
        t.start()
        threads.append(t)

    interrupted = False
    try:
        for t in threads:
            while t.is_alive():
                t.join(timeout=0.5)
    except KeyboardInterrupt:
        interrupted = True
        stop_event.set()
        print("\n⚠️  收到 Ctrl+C，正在停止清理流程...")
        for t in threads:
            t.join(timeout=1)
    finally:
        stop_event.set()
        for t in threads:
            if t.is_alive():
                t.join(timeout=0.5)

    if interrupted:
        return {
            "input_rows": len(df),
            "kept_rows": 0,
            "removed_rows": 0,
            "removed_indexes": [],
            "results": results,
            "errors": errors,
            "xlsx_output": output_path,
            "csv_output": output_path.with_suffix(".csv"),
            "interrupted": True,
        }

    kept_indices = [idx for idx in df.index if results.get(idx, {}).get("keep")]
    removed = [idx for idx in df.index if idx not in kept_indices]
    cleaned_df = df.loc[kept_indices].copy()

    cleaned_df.to_excel(output_path, index=False, engine="openpyxl", sheet_name=sheet_name)
    _style_output_xlsx(output_path)

    csv_output = output_path.with_suffix(".csv")
    cleaned_df.to_csv(csv_output, index=False, encoding="utf-8-sig")

    return {
        "input_rows": len(df),
        "kept_rows": len(cleaned_df),
        "removed_rows": len(removed),
        "removed_indexes": removed,
        "results": results,
        "errors": errors,
        "xlsx_output": output_path,
        "csv_output": csv_output,
        "interrupted": False,
    }


def main():
    parser = argparse.ArgumentParser(description="Validate Review URL rows using live Glassdoor pages and remove mismatches.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Input XLSX/CSV file")
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output XLSX file (CSV uses the same stem). Default: <input>_cleaned.xlsx",
    )
    parser.add_argument(
        "--ports",
        nargs="+",
        type=int,
        default=DEFAULT_PORTS,
        help="Chrome debug ports to use (default: 9225 9226 9227)",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    output_path = args.output.resolve() if args.output else input_path.with_name(f"{input_path.stem}_cleaned.xlsx")

    print(f"Input : {input_path}")
    print(f"Output: {output_path}")
    print(f"Ports : {args.ports}")

    started = time.time()
    try:
        report = clean_workbook(input_path, output_path, args.ports)
    except KeyboardInterrupt:
        print("\n⚠️  已中斷")
        return
    elapsed = time.time() - started

    if report.get("interrupted"):
        print("\n⚠️  已中斷，未輸出清理後檔案")
        return

    print("")
    print(f"Kept   : {report['kept_rows']} / {report['input_rows']}")
    print(f"Removed: {report['removed_rows']}")
    if report["errors"]:
        print("Port errors:")
        for err in report["errors"]:
            print(f"  - {err}")

    if report["removed_rows"]:
        print("Removed rows:")
        for idx in report["removed_indexes"][:30]:
            info = report["results"].get(idx, {})
            row_no = idx + 2  # Excel row number (header row is 1)
            print(
                f"  - row {row_no}: {info.get('reason', 'invalid')} | "
                f"{info.get('company', '')} | overall={info.get('overall', '')} | "
                f"{info.get('baseline', '')} | {info.get('url', '')}"
            )
        if report["removed_rows"] > 30:
            print(f"  ... and {report['removed_rows'] - 30} more")

    print("")
    print(f"Wrote: {report['xlsx_output']}")
    print(f"Wrote: {report['csv_output']}")
    print(f"Done in {elapsed:.1f}s")

    if report["removed_rows"] == 0:
        print("No rows were removed.")


if __name__ == "__main__":
    main()
