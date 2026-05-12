import pandas as pd
import json
import glob
import os

# 讀現有 Excel
df = pd.read_excel('data/glassdoor_ratings.xlsx')
print(f"現有 {len(df)} 筆，欄位：{list(df.columns)}")
print(df[['Company', 'Baseline Location']].head(5).to_string())

# 建立 (company, baseline_location) -> url mapping
url_map = {}

# 1. asus_locations.json（explore 格式）
with open('data/asus_locations.json', encoding='utf-8') as f:
    asus = json.load(f)
for entry in asus:
    if entry.get('url'):
        key = ('ASUS', entry['location'])
        url_map[key] = entry['url']

# 2. *_matched.json（match 格式）
for path in glob.glob('data/*_matched.json'):
    with open(path, encoding='utf-8') as f:
        entries = json.load(f)
    for entry in entries:
        if entry.get('url') and entry.get('status') == 'found':
            key = (entry['company'], entry['baseline_location'])
            url_map[key] = entry['url']

print(f"\nURL mapping：{len(url_map)} 筆")

# 匹配
def lookup_url(row):
    company_base = row['Company'].split(' - ')[0].strip()
    baseline = row['Baseline Location']
    return url_map.get((company_base, baseline))

df['Review URL'] = df.apply(lookup_url, axis=1)
matched = df['Review URL'].notna().sum()
print(f"\n成功對應 URL：{matched}/{len(df)} 筆")
print(df[df['Review URL'].isna()][['Company', 'Baseline Location']].to_string())

# 欄位排序：Review URL 放在 Actual City 後面
cols = list(df.columns)
cols.remove('Review URL')
insert_at = cols.index('Actual City') + 1
cols.insert(insert_at, 'Review URL')
df = df[cols]

# 存回 Excel
output = 'data/glassdoor_ratings.xlsx'
df.to_excel(output, index=False, engine='openpyxl')

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
wb = openpyxl.load_workbook(output)
ws = wb.active

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Review URL 欄寬加寬
url_col_idx = cols.index('Review URL') + 1
url_col_letter = openpyxl.utils.get_column_letter(url_col_idx)
ws.column_dimensions[url_col_letter].width = 60

for i, col in enumerate(ws.columns, 1):
    letter = openpyxl.utils.get_column_letter(i)
    if letter != url_col_letter:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[letter].width = min(max_len + 2, 30)

wb.save(output)

# 同步更新 CSV
df.to_csv('data/glassdoor_ratings.csv', index=False, encoding='utf-8-sig')
print(f"✓ 已存回 {output} 和 data/glassdoor_ratings.csv")
