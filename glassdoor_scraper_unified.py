"""
Glassdoor 爬蟲工具 - 整合版
支援兩種模式：
1. 自動模式：自動啟動 Chrome（可能遇到反爬蟲）
2. 手動模式：連接到已登入的 Chrome（推薦）
"""

import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import re
import os
import json
from concurrent.futures import ThreadPoolExecutor, as_completed

class GlassdoorScraper:
    def __init__(self, mode='manual', headless=False, chrome_debugger_port=9222, _log_fn=None):
        """
        初始化 Glassdoor 爬蟲
        
        Args:
            mode: 'manual' 或 'auto'
                - 'manual': 連接到已登入的 Chrome（推薦，避免反爬蟲）
                - 'auto': 自動啟動新的 Chrome（可能遇到人機驗證）
            headless: 僅在 auto 模式下有效，是否隱藏瀏覽器
            chrome_debugger_port: 僅在 manual 模式下有效，Chrome 調試端口
            _log_fn: 可選的 logging 函式 (msg, end) → 供平行模式各 port 獨立記錄
        """
        self.mode = mode
        self._port = chrome_debugger_port
        self._log = _log_fn if _log_fn else lambda msg='', end='\n': print(msg, end=end)
        chrome_options = Options()
        
        if mode == 'manual':
            chrome_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{chrome_debugger_port}")
            self._log(f"📌 使用手動模式：連接到端口 {chrome_debugger_port} 的 Chrome")
        else:
            if headless:
                chrome_options.add_argument('--headless')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            self._log(f"📌 使用自動模式：啟動新的 Chrome（headless={headless}）")
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.wait = WebDriverWait(self.driver, 15)

    def _is_blocked(self):
        """Check if the current page is a Cloudflare / 'Humans only' block page."""
        try:
            page_source = self.driver.page_source.lower()
            # Only match actual challenge page indicators, not generic CDN references
            # Note: 'just a moment' removed from body check (too many false positives);
            # it is caught by the title check below instead.
            indicators = ['humans only', 'checking your browser',
                          'cf-challenge-running', 'verify you are human',
                          'cf-turnstile', 'challenge-platform']
            # Additional check: page title is typically "Just a moment..." on block pages
            title_blocked = self.driver.title.strip().lower() in ['just a moment...', 'attention required']
            return title_blocked or any(ind in page_source for ind in indicators)
        except Exception:
            return False

    def _wait_if_blocked(self, context='', timeout=300, auto_refresh_interval=15):
        """If Cloudflare block is detected, automatically refresh the page periodically
        until unblocked or timeout is reached.

        Args:
            context: descriptive label for logging
            timeout: total seconds before giving up on this port
            auto_refresh_interval: seconds between automatic page refreshes (default 15s)

        Returns:
            True if page recovered (or was never blocked), False if timed out.
        """
        if not self._is_blocked():
            return True
        port = self._port
        msg = f"[BLOCKED] Port {port} 被 Cloudflare 攔截（{context}），將每 {auto_refresh_interval}s 自動刷新頁面..."
        self._log(msg)
        print(msg, flush=True)
        wait_count = 0
        refresh_count = 0
        while True:
            time.sleep(5)
            wait_count += 1
            elapsed = wait_count * 5

            try:
                # Auto-refresh every auto_refresh_interval seconds
                if elapsed % auto_refresh_interval == 0:
                    refresh_count += 1
                    refresh_msg = f"[BLOCKED] Port {port} 自動刷新頁面（第 {refresh_count} 次，已等待 {elapsed}s）..."
                    self._log(refresh_msg)
                    print(refresh_msg, flush=True)
                    self.driver.refresh()
                    # Give the page a moment to start loading after refresh
                    time.sleep(3)

                if not self._is_blocked():
                    msg = f"[BLOCKED] Port {port} 已恢復（刷新 {refresh_count} 次後），繼續抓取"
                    self._log(msg)
                    print(msg, flush=True)
                    return True
            except Exception:
                pass

            if elapsed >= timeout:
                msg = f"[BLOCKED] Port {port} 超過 {timeout}s 未恢復，暫停此 port，任務轉移給其他 port"
                self._log(msg)
                print(msg, flush=True)
                return False

    def extract_rating_data(self, url, company_name):
        """
        從 Glassdoor 頁面提取評分數據
        
        Args:
            url: Glassdoor 評論頁面 URL
            company_name: 公司名稱（用於標識）
            
        Returns:
            dict: 包含各項評分的字典
        """
        self._log(f"正在抓取: {company_name}")
        self._log(f"URL: {url}")
        _t0 = time.time()
        
        try:
            self.driver.get(url)
            # 給 Cloudflare JS challenge 短暫時間自動通過（最多 5 秒）
            for _ in range(5):
                if self.driver.title.strip().lower() not in ['just a moment...', 'attention required', '']:
                    break
                time.sleep(1)

            # 偵測 Cloudflare / "Humans only" 攔截頁面
            if not self._wait_if_blocked(company_name):
                # Timed out — return sentinel so caller can stop this port
                return 'BLOCKED_TIMEOUT'

            # 等 Overall 評分出現（最多 8 秒），取代固定 sleep(5)
            data = {
                'Company': company_name,
                'Overall': None,
                'Recommend': None,
                'CEO Approval': None,
                'Total Reviews': None,
                'Diversity & Inclusion': None,
                'Work/Life Balance': None,
                'Compensation and Benefits': None,
                'Culture & Values': None,
                'Career Opportunities': None,
                'Senior Management': None
            }

            # 提取整體評分
            try:
                overall_elem = WebDriverWait(self.driver, 8).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'p[data-size-variant="lg"]'))
                )
                overall_rating = overall_elem.text
                data['Overall'] = float(overall_rating)
                self._log(f"  ✓ Overall: {overall_rating}")
            except Exception as e:
                self._log(f"  ⚠ 無法找到整體評分")

            # 滾動頁面以觸發動態載入
            self.driver.execute_script("window.scrollTo(0, 500);")
            time.sleep(1)
            
            # 提取推薦比例
            try:
                recommend_elem = self.driver.find_element(By.CSS_SELECTOR, 'p[data-test="recommendToFriend"]')
                recommend_text = recommend_elem.text
                match = re.search(r'(\d+)%', recommend_text)
                if match:
                    data['Recommend'] = f"{match.group(1)}%"
                    self._log(f"  ✓ Recommend: {data['Recommend']}")
            except Exception as e:
                self._log(f"  ⚠ 無法找到推薦比例")
            
            # 提取 CEO 支持率
            try:
                ceo_elem = self.driver.find_element(By.CSS_SELECTOR, 'p[class*="ceoApproval"]')
                ceo_match = re.search(r'(\d+)%', ceo_elem.text)
                if ceo_match:
                    data['CEO Approval'] = f"{ceo_match.group(1)}%"
                    self._log(f"  ✓ CEO Approval: {data['CEO Approval']}")
            except Exception:
                self._log(f"  ⚠ 無法找到 CEO 支持率")

            # 提取評論總數
            try:
                review_count_elem = self.driver.find_element(By.CSS_SELECTOR, '.ReviewOverview_count__xexV9')
                review_text = review_count_elem.get_attribute('textContent')
                match = re.search(r'([\d,]+)', review_text)
                if match:
                    data['Total Reviews'] = match.group(1)
                    self._log(f"  ✓ Total Reviews: {data['Total Reviews']}")
            except Exception as e:
                self._log(f"  ⚠ 無法找到評論總數")
            
            # 提取各項評分
            try:
                rating_items = self.driver.find_elements(By.CSS_SELECTOR, '.RatingsByCategory_ratingItem__4EMZd')
                
                for item in rating_items:
                    try:
                        rating_value = item.find_element(By.CSS_SELECTOR, '.RatingsByCategory_rating__T0v8N').text
                        category_label = item.find_element(By.CSS_SELECTOR, '.RatingsByCategory_ratingLabel__o3Me9').text
                        
                        if 'Diversity' in category_label or 'inclusion' in category_label:
                            data['Diversity & Inclusion'] = float(rating_value)
                            self._log(f"  ✓ Diversity & Inclusion: {rating_value}")
                        elif 'Work/Life' in category_label or 'Work-Life' in category_label:
                            data['Work/Life Balance'] = float(rating_value)
                            self._log(f"  ✓ Work/Life Balance: {rating_value}")
                        elif 'Compensation' in category_label or 'benefits' in category_label:
                            data['Compensation and Benefits'] = float(rating_value)
                            self._log(f"  ✓ Compensation and Benefits: {rating_value}")
                        elif 'Culture' in category_label or 'values' in category_label:
                            data['Culture & Values'] = float(rating_value)
                            self._log(f"  ✓ Culture & Values: {rating_value}")
                        elif 'Career' in category_label or 'opportunities' in category_label:
                            data['Career Opportunities'] = float(rating_value)
                            self._log(f"  ✓ Career Opportunities: {rating_value}")
                        elif 'Senior' in category_label or 'management' in category_label:
                            data['Senior Management'] = float(rating_value)
                            self._log(f"  ✓ Senior Management: {rating_value}")
                    except Exception as e:
                        continue
                        
            except Exception as e:
                self._log(f"  ⚠ 提取評分時發生錯誤: {str(e)}")
            
            # 檢查缺失的數據（排除由呼叫端填入的 metadata 欄位）
            _meta_cols = {'Baseline Location', 'Actual City', 'Country'}
            missing_fields = [k for k, v in data.items() if k not in _meta_cols and k != 'Company' and v is None]
            if missing_fields:
                self._log(f"  ⚠ 缺失欄位: {', '.join(missing_fields)}")
            
            self._log(f"✓ 成功抓取 {company_name} ({time.time() - _t0:.1f}s)")
            return data
            
        except Exception as e:
            self._log(f"✗ 抓取 {company_name} 時發生錯誤: {str(e)} ({time.time() - _t0:.1f}s)")
            return None
    
    def scrape_multiple_companies(self, company_urls):
        """
        抓取多個公司的數據
        
        Args:
            company_urls: 字典，格式為 {公司名稱: URL}
            
        Returns:
            list: 包含所有公司數據的列表
        """
        all_data = []
        
        for company_name, url in company_urls.items():
            data = self.extract_rating_data(url, company_name)
            if data:
                all_data.append(data)
            time.sleep(3)
            self._log()
        
        return all_data

    def scrape_from_matched_json(self, json_files, _progress_fn=None):
        """
        從 company_finder.py 產生的 URL list JSON 檔案抓取評分數據。
        支援 office/country/scan 三種類型。

        Args:
            json_files: list of str，URL list JSON 路徑
            _progress_fn: callable(entry_name)，每筆 entry 完成後回報進度

        Returns:
            list of dict
        """
        import json
        all_data = []
        total_start = time.time()
        count = 0

        for json_file in json_files:
            # Detect source mode from filename
            basename = os.path.basename(json_file)
            if basename.endswith('_country.json'):
                source_mode = 'country'
            elif basename.endswith('_scan.json'):
                source_mode = 'scan'
            elif basename.endswith('_city.json'):
                source_mode = 'city'
            elif basename.endswith('_office.json'):
                source_mode = 'office'
            else:
                source_mode = 'unknown'

            with open(json_file, encoding='utf-8') as f:
                entries = json.load(f)

            for entry in entries:
                if entry['status'] != 'found' or not entry.get('url'):
                    continue
                company = entry['company']
                baseline_loc = entry.get('baseline_location') or entry.get('country', 'Unknown')
                actual_city = entry.get('matched_city') or baseline_loc
                url = entry['url']

                country = entry.get('baseline_country') or entry.get('country')
                display_name = f"{company} - {baseline_loc}"
                data = self.extract_rating_data(url, display_name)
                if data == 'BLOCKED_TIMEOUT':
                    return 'BLOCKED_TIMEOUT'
                sleep_sec = 0.5
                if data:
                    data['Baseline Location'] = baseline_loc
                    data['Actual City'] = actual_city
                    data['Country'] = country
                    data['Review URL'] = url
                    data['Source Mode'] = source_mode  # city / country / scan
                    all_data.append(data)
                    # 成功且有關鍵欄位才等較久，否則短等待
                    if data.get('Overall') is not None:
                        sleep_sec = 1.5
                count += 1
                # 回報進度
                if _progress_fn:
                    _progress_fn(display_name)
                time.sleep(sleep_sec)
                self._log()

        total = time.time() - total_start
        if count:
            self._log(f"⏱ 總計 {count} 筆，耗時 {total:.0f}s，平均 {total/count:.1f}s/筆")
        return all_data
    
    def scrape_from_baseline_json(self, json_file, company_name='ASUS', location_filter=None, _progress_fn=None):
        """
        讀取 office 模式產生的 *_office.json（如 asus_office.json）並抓取評分。

        Args:
            json_file: str，office locations json 路徑
            company_name: str，公司名稱
            location_filter: str or None，若指定則只處理該地區
            _progress_fn: callable(entry_name)，每筆 entry 完成後回報進度

        Returns:
            list of dict
        """
        import json
        all_data = []

        with open(json_file, encoding='utf-8') as f:
            entries = json.load(f)

        for entry in entries:
            if entry.get('status') != 'found' or not entry.get('url'):
                continue
            baseline_loc = entry['location']
            if location_filter and baseline_loc != location_filter:
                continue
            country = entry.get('country')
            url = entry['url']

            display_name = f"{company_name} - {baseline_loc}"
            data = self.extract_rating_data(url, display_name)
            if data == 'BLOCKED_TIMEOUT':
                return 'BLOCKED_TIMEOUT'
            sleep_sec = 0.5
            if data:
                data['Baseline Location'] = baseline_loc
                data['Actual City'] = baseline_loc
                data['Country'] = country
                data['Review URL'] = url
                data['Source Mode'] = 'office'
                all_data.append(data)
                if data.get('Overall') is not None:
                    sleep_sec = 1.5
            # 回報進度
            if _progress_fn:
                _progress_fn(display_name)
            time.sleep(sleep_sec)
            self._log()

        return all_data

    def save_to_excel(self, data, output_file='glassdoor_ratings.xlsx'):
        """
        將數據保存為 Excel 文件
        
        Args:
            data: 數據列表
            output_file: 輸出文件名
        """
        if not data:
            print("沒有數據可以保存")
            return
        
        df = pd.DataFrame(data)

        # 調整列順序（相容有無 Location 欄位）
        # Source Mode 放前面方便識別
        base_cols = ['Company', 'Source Mode', 'Baseline Location', 'Country', 'Actual City', 'Review URL',
                     'Overall', 'Recommend', 'CEO Approval', 'Total Reviews',
                     'Diversity & Inclusion', 'Work/Life Balance', 'Compensation and Benefits',
                     'Culture & Values', 'Career Opportunities', 'Senior Management']
        columns_order = [c for c in base_cols if c in df.columns]
        df = df[columns_order]
        
        # 保存為 Excel
        os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
        df.to_excel(output_file, index=False, engine='openpyxl')
        print(f"\n✓ 數據已保存至: {output_file}")  # save_to_excel 由主 thread 呼叫，用 print
        
        # 美化 Excel
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = load_workbook(output_file)
        ws = wb.active
        
        # 設置標題樣式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 設置列寬
        col_widths = {'A': 30, 'B': 15, 'C': 22, 'D': 22, 'E': 10, 'F': 12,
                      'G': 14, 'H': 20, 'I': 18, 'J': 24, 'K': 16,
                      'L': 22, 'M': 18}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Color coding for Source Mode column
        mode_colors = {
            'city': 'CCE5FF',               # Light blue
            'country': 'D4EDDA',             # Light green
            'office': 'E8D5F5',              # Light purple
            'scan': 'FFF3CD',                # Light orange
            'unknown': 'F8D7DA',             # Light red
            'city+country_fb': 'B8D4F0',     # Darker blue (city with country fallback)
            'city+office_fb': 'C8C8F0',      # Blue-purple (city with office fallback)
            'office+country_fb': 'D4C8F5',   # Darker purple (office with country fallback)
            'office+office_fb': 'C0A8E8',    # Deep purple (office with office fallback)
        }

        # Find Source Mode column index
        source_mode_col = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == 'Source Mode':
                source_mode_col = idx
                break

        # Apply colors to Source Mode column (prefix match for fallback variants)
        if source_mode_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[source_mode_col - 1]  # 0-indexed
                mode = cell.value or ''
                color = mode_colors.get(mode) or mode_colors.get(mode.split('+')[0])
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # 居中對齊數據
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        wb.save(output_file)
        print(f"✓ Excel 格式已美化")

        # 同步輸出 CSV
        csv_file = os.path.splitext(output_file)[0] + '.csv'
        df.to_csv(csv_file, index=False, encoding='utf-8-sig')
        print(f"✓ CSV 已保存至: {csv_file}")  # save_to_excel 由主 thread 呼叫，用 print
    
    def close(self):
        """關閉瀏覽器"""
        if self.mode == 'auto':
            self.driver.quit()
        else:
            # 手動模式不關閉瀏覽器
            pass


def _worker(port, task_queue, mode, headless, log_path=None, _progress=None):
    """
    單一執行緒的工作函式，從共享 task_queue 取任務。
    若被 Cloudflare 攔截超時，剩餘任務留在 queue 給其他 port。

    Args:
        port: Chrome debug port
        task_queue: queue.Queue，共享任務佇列
        mode: scraper mode
        headless: bool
        log_path: 此 port 專屬的 log 路徑，None 則不寫檔
        _progress: dict，用於回報進度給主執行緒

    Returns:
        list of dict
    """
    import io
    import threading
    import queue

    buf = io.StringIO()
    _lock = threading.Lock()
    _log_file = None

    if log_path:
        os.makedirs(os.path.dirname(log_path), exist_ok=True)
        _log_file = open(log_path, 'w', encoding='utf-8')

    def _log(msg='', end='\n'):
        s = str(msg) + end
        with _lock:
            buf.write(s)
            if _log_file:
                _log_file.write(s)
                _log_file.flush()

    scraper = GlassdoorScraper(mode=mode, headless=headless,
                               chrome_debugger_port=port, _log_fn=_log)
    results = []
    blocked_out = False
    try:
        while True:
            try:
                task = task_queue.get_nowait()
            except queue.Empty:
                break

            task_name = task.get('name') or task.get('company_name', task.get('file', 'unknown'))
            if task.get('location_filter'):
                task_name += f" [{task['location_filter']}]"

            def _progress_fn(entry_name):
                if _progress is not None:
                    with threading.Lock():
                        _progress[port]['completed'] += 1
                        _progress[port]['current'] = entry_name

            if task['type'] == 'baseline':
                data = scraper.scrape_from_baseline_json(
                    task['file'], task['company_name'],
                    location_filter=task.get('location_filter'),
                    _progress_fn=_progress_fn
                )
            elif task['type'] == 'matched_entry':
                entry = task['entry']
                src_mode = task['source_mode']
                # 'company' field only exists in country/city entries; derive from task name for office/scan
                company = entry.get('company') or task['name'].split(' - ')[0]
                baseline_loc = entry.get('baseline_location') or entry.get('location') or entry.get('country', 'Unknown')
                actual_city = entry.get('matched_city') or baseline_loc
                url = entry['url']
                country = entry.get('baseline_country') or entry.get('country')
                display_name = f"{company} - {baseline_loc}"
                raw = scraper.extract_rating_data(url, display_name)
                if raw == 'BLOCKED_TIMEOUT':
                    data = 'BLOCKED_TIMEOUT'
                else:
                    data = None  # will be set below; initialize to avoid UnboundLocalError
                    sleep_sec = 0.5
                    # --- Country fallback: if page has no rating data, try fallback URL ---
                    fallback_info = task.get('fallback')
                    if raw and raw.get('Overall') is None and fallback_info:
                        fb_url = fallback_info['url']
                        fb_mode = fallback_info['mode']
                        _log(f"  ↩ {baseline_loc} 無評分資料，嘗試 fallback ({fb_mode}): {fb_url}")
                        time.sleep(0.5)
                        fb_raw = scraper.extract_rating_data(fb_url, f"{display_name} [fb]")
                        if fb_raw == 'BLOCKED_TIMEOUT':
                            data = 'BLOCKED_TIMEOUT'
                        else:
                            if fb_raw and fb_raw.get('Overall') is not None:
                                _log(f"  ✓ fallback 成功，Overall={fb_raw['Overall']}")
                                raw = fb_raw
                                src_mode = f"{src_mode}+{fb_mode}_fb"
                                url = fb_url
                            else:
                                _log(f"  ⚠ fallback 也無評分資料，保留空值")
                    if data != 'BLOCKED_TIMEOUT':
                        if raw:
                            raw['Baseline Location'] = baseline_loc
                            raw['Actual City'] = actual_city
                            raw['Country'] = country
                            raw['Review URL'] = url
                            raw['Source Mode'] = src_mode
                            if raw.get('Overall') is not None:
                                sleep_sec = 1.5
                            data = [raw]
                        else:
                            data = []
                        _progress_fn(display_name)
                        time.sleep(sleep_sec)
            else:
                data = []

            # Check if scraper hit a block timeout (sentinel returned inside results)
            if data == 'BLOCKED_TIMEOUT' or (isinstance(data, list) and 'BLOCKED_TIMEOUT' in data):
                # Put this task back in the queue for other ports
                task_queue.put(task)
                _log(f"[BLOCKED] Port {port} 任務 {task_name} 已放回佇列")
                print(f"[BLOCKED] Port {port} 任務 {task_name} 已放回佇列", flush=True)
                blocked_out = True
                break

            if isinstance(data, list):
                results += data

            task_queue.task_done()

    finally:
        scraper.close()
        if _log_file:
            _log_file.close()
        if _progress is not None:
            with threading.Lock():
                _progress[port]['finished'] = True
                if blocked_out:
                    _progress[port]['blocked'] = True
    return results


def parallel_scrape(ports, matched_files, include_baseline, baseline_file,
                    mode='manual', headless=False, log_ts=None):
    """
    使用共享任務佇列讓多個 Chrome port 平行執行。
    被 Cloudflare 攔截超時的 port 會停止，剩餘任務由其他 port 接手。

    Returns:
        list of dict
    """
    import threading
    import queue as queue_mod

    tasks_all = []

    # Skip baseline if asus_office.json or asus_country.json is already in matched_files
    # (avoids scraping ASUS twice)
    asus_in_matched = any(
        os.path.basename(f) in ('asus_office.json', 'asus_country.json', 'asus_city.json')
        for f in matched_files
    )
    if include_baseline and asus_in_matched:
        print("[INFO] asus_office.json/asus_country.json/asus_city.json 已在 matched files 中，略過 baseline 避免重複")
        include_baseline = False

    if include_baseline and os.path.exists(baseline_file):
        with open(baseline_file, encoding='utf-8') as _f:
            _entries = json.load(_f)
        for _e in _entries:
            if _e.get('status') == 'found' and _e.get('url'):
                tasks_all.append({
                    'type': 'baseline',
                    'file': baseline_file,
                    'company_name': 'ASUS',
                    'location_filter': _e['location'],
                })

    # --- Build country-level fallback lookup table ---
    # Scans ALL *_country.json and *_office.json in the data/ directory (not just
    # matched_files), so fallback works even when the user only selected city files.
    # Maps (company_lower, country_lower) -> {'url': ..., 'mode': 'country'|'office'}
    # Priority: country entry > office entry with matching country > office Global entry
    _country_fallback: dict = {}  # (company_lower, country_lower) -> url
    _global_fallback: dict = {}   # company_lower -> url (Global office/country entry)

    # Collect all country/office JSON files from the data/ directory.
    # baseline_file is typically 'data/asus_office.json', so its parent is already 'data/'.
    _data_dir = os.path.dirname(os.path.abspath(baseline_file))
    if not os.path.isdir(_data_dir):
        _data_dir = 'data'
    _fallback_files = []
    if os.path.isdir(_data_dir):
        for _fn in os.listdir(_data_dir):
            if _fn.endswith('_country.json') or _fn.endswith('_office.json'):
                _fallback_files.append(os.path.join(_data_dir, _fn))
    # Also include baseline_file itself (asus_office.json)
    if os.path.exists(baseline_file) and baseline_file not in _fallback_files:
        _fallback_files.append(baseline_file)

    for _fb_file in _fallback_files:
        _fb_basename = os.path.basename(_fb_file)
        _fb_company = _fb_basename.replace('_office.json', '').replace('_country.json', '').replace('_', ' ').title()
        try:
            with open(_fb_file, encoding='utf-8') as _fh:
                _entries_fb = json.load(_fh)
        except Exception:
            continue
        for _e in _entries_fb:
            if _e.get('status') != 'found' or not _e.get('url'):
                continue
            _c = (_e.get('company') or _fb_company).lower()
            if _fb_basename.endswith('_country.json'):
                _country = (_e.get('baseline_country') or _e.get('matched_city') or '').lower()
                if _country and _country not in ('global', ''):
                    _key = (_c, _country)
                    if _key not in _country_fallback:
                        _country_fallback[_key] = {'url': _e['url'], 'mode': 'country'}
                if (_e.get('baseline_location') or '').lower() in ('global', '') or _country in ('global', ''):
                    if _c not in _global_fallback:
                        _global_fallback[_c] = {'url': _e['url'], 'mode': 'country'}
            elif _fb_basename.endswith('_office.json'):
                _loc = (_e.get('location') or '').lower()
                _cntry = (_e.get('country') or '').lower()
                if _loc == 'global' or _cntry == 'global':
                    if _c not in _global_fallback:
                        _global_fallback[_c] = {'url': _e['url'], 'mode': 'office'}
                elif _cntry and _cntry != 'global':
                    _key = (_c, _cntry)
                    if _key not in _country_fallback:
                        _country_fallback[_key] = {'url': _e['url'], 'mode': 'office'}

    def _get_fallback(company: str, country: str):
        """Return fallback dict {url, mode} or None."""
        _c = company.lower()
        _cntry = (country or '').lower()
        if _cntry and _cntry not in ('global', 'unknown', ''):
            fb = _country_fallback.get((_c, _cntry))
            if fb:
                return fb
        return _global_fallback.get(_c)

    for f in matched_files:
        basename = os.path.basename(f)
        if basename.endswith('_country.json'):
            src_mode = 'country'
        elif basename.endswith('_scan.json'):
            src_mode = 'scan'
        elif basename.endswith('_city.json'):
            src_mode = 'city'
        elif basename.endswith('_office.json'):
            src_mode = 'office'
        else:
            src_mode = 'unknown'
        # Derive company name from filename for office/scan modes (no 'company' field in entries)
        file_company = basename.replace('_office.json', '').replace('_scan.json', '').replace('_city.json', '').replace('_country.json', '').replace('_', ' ').title()
        with open(f, encoding='utf-8') as _fh:
            _entries = json.load(_fh)
        for _e in _entries:
            if _e.get('status') != 'found' or not _e.get('url'):
                continue
            _company = _e.get('company') or file_company
            _baseline_loc = _e.get('baseline_location') or _e.get('location') or _e.get('country', 'Unknown')
            _entry_country = _e.get('baseline_country') or _e.get('country') or ''
            # Attach fallback only for city/office modes (country mode already is country-level)
            _fallback = None
            if src_mode in ('city', 'office'):
                _fb = _get_fallback(_company, _entry_country)
                # Only use fallback if it's a different URL from the entry itself
                if _fb and _fb['url'] != _e['url']:
                    _fallback = _fb
            tasks_all.append({
                'type': 'matched_entry',
                'entry': _e,
                'source_mode': src_mode,
                'name': f"{_company} - {_baseline_loc}",
                'fallback': _fallback,  # {'url': ..., 'mode': ...} or None
            })

    if not tasks_all:
        return []

    # Shared task queue
    task_queue = queue_mod.Queue()
    for task in tasks_all:
        task_queue.put(task)

    ts = log_ts or ''
    port_logs = {port: f'logs/run_{ts}_port{port}.txt' for port in ports}

    # 進度追蹤
    _progress = {}
    for port in ports:
        _progress[port] = {'completed': 0, 'current': 'init', 'finished': False, 'blocked': False}

    # Each task is now a single entry — total is simply len(tasks_all)
    total_tasks = len(tasks_all)

    total_start = time.time()
    all_data = []
    port_counts = {}

    _last_printed = 0

    def _print_progress(force=False):
        """Print simple text progress for dashboard parsing."""
        nonlocal _last_printed
        total_completed = sum(p['completed'] for p in _progress.values())
        if total_completed == _last_printed and not force:
            return
        _last_printed = total_completed
        parts = []
        for port in sorted(_progress.keys()):
            p = _progress[port]
            if p.get('blocked'):
                parts.append(f"P{port}:BLOCKED({p['completed']})")
            elif p['finished']:
                parts.append(f"P{port}:OK({p['completed']})")
            else:
                parts.append(f"P{port}:{p['completed']}")
        print(f"[PROGRESS] {total_completed}/{total_tasks} | {' | '.join(parts)}", flush=True)

    executor = ThreadPoolExecutor(max_workers=len(ports))
    futures = {
        executor.submit(_worker, port, task_queue, mode, headless, port_logs[port], _progress): port
        for port in ports
    }

    try:
        while futures:
            done_futures = []
            for f in list(futures):
                if f.done():
                    port = futures[f]
                    try:
                        data = f.result()
                        all_data += data
                        port_counts[port] = len(data)
                        blocked = _progress[port].get('blocked', False)
                        status = "BLOCKED-OUT" if blocked else "completed"
                        print(f"[DONE] Port {port} {status} {len(data)} entries  log: {port_logs[port]}", flush=True)
                    except Exception as e:
                        print(f"[ERROR] Port {port} error: {e}", flush=True)
                    done_futures.append(f)
                    _progress[port]['finished'] = True

            for f in done_futures:
                del futures[f]

            # If there are still tasks in queue but all workers are done, some ports blocked out.
            # Check if we should report remaining tasks.
            if not futures and not task_queue.empty():
                remaining = task_queue.qsize()
                print(f"[WARN] {remaining} tasks remaining in queue but all ports stopped (blocked). These tasks were not completed.", flush=True)

            _print_progress()

            if futures:
                time.sleep(3)

    except KeyboardInterrupt:
        print("\n[WARN] Ctrl+C received, stopping workers...", flush=True)
        for f in futures:
            f.cancel()
        executor.shutdown(wait=False, cancel_futures=True)
        raise

    elapsed = time.time() - total_start
    print(f"\n[DONE] Parallel scraping completed in {elapsed:.0f}s, {len(all_data)} entries total", flush=True)

    # 將各 port log 合並追加到主 log
    import sys as _sys
    main_log = getattr(getattr(_sys.stdout, '_file', None), 'name', None)
    if main_log:
        try:
            with open(main_log, 'a', encoding='utf-8') as mf:
                mf.write(f"\n{'='*60}\n")
                mf.write(f"Per-Port Detailed Logs\n")
                mf.write(f"{'='*60}\n")
                for port in ports:
                    plog = port_logs[port]
                    if os.path.exists(plog):
                        mf.write(f"\n--- port {port} ({port_counts.get(port, 0)} 筆) ---\n")
                        with open(plog, encoding='utf-8') as pf:
                            mf.write(pf.read())
        except Exception:
            pass

    return all_data


class TeeLogger:
    """將 stdout 同時輸出到終端和檔案"""
    def __init__(self, filepath):
        import sys
        self._terminal = sys.stdout
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        self._file = open(filepath, 'w', encoding='utf-8')
        import sys as _sys
        _sys.stdout = self

    def write(self, message):
        try:
            self._terminal.write(message)
        except UnicodeEncodeError:
            self._terminal.write(message.encode(self._terminal.encoding or 'utf-8', errors='replace').decode(self._terminal.encoding or 'utf-8', errors='replace'))
        self._file.write(message)

    def flush(self):
        self._terminal.flush()
        self._file.flush()

    def restore(self):
        import sys
        sys.stdout = self._terminal
        self._file.close()


def save_run_report(data, elapsed, log_txt_path, json_path):
    """產出 JSON 摘要報告"""
    from datetime import datetime
    import collections

    company_stats = collections.defaultdict(lambda: {'success': 0, 'failed': 0, 'missing': []})
    for row in data:
        company = row.get('Company', 'Unknown').split(' - ')[0]
        company_stats[company]['success'] += 1
        missing = [k for k, v in row.items()
                   if v is None and k not in ('Company', 'Baseline Location', 'Actual City', 'Country')]
        if missing:
            company_stats[company]['missing'].append({'location': row.get('Baseline Location'), 'fields': missing})

    report = {
        'run_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'elapsed_seconds': round(elapsed, 1),
        'elapsed_human': f"{int(elapsed//60)}m {int(elapsed%60)}s",
        'total_rows': len(data),
        'log_file': log_txt_path,
        'companies': {k: dict(v) for k, v in company_stats.items()},
    }

    os.makedirs(os.path.dirname(json_path), exist_ok=True)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"✓ JSON 摘要已保存至: {json_path}")


def load_config():
    """載入配置文件"""
    try:
        # 嘗試導入 config.py
        import config
        return config
    except ImportError:
        # 如果沒有 config.py，嘗試 config_example.py
        try:
            import config_example as config
            print("⚠ 使用 config_example.py，建議複製為 config.py 並修改")
            return config
        except ImportError:
            print("✗ 找不到配置文件")
            return None


def main():
    import glob
    import argparse
    from datetime import datetime

    parser = argparse.ArgumentParser(description='Glassdoor Scraper')
    parser.add_argument('--mode', choices=['manual', 'auto'], default=None, help='Chrome connection mode (manual=attach to existing, auto=launch new)')
    parser.add_argument('--task', choices=['matched', 'baseline'], default=None, help='Scraping task type')
    parser.add_argument('--ports', default=None, help='Comma-separated Chrome debug ports')
    parser.add_argument('--source-mode', choices=['all', 'office', 'city', 'country', 'scan'], default='all', help='Filter matched files by source mode (office/city/country/scan)')
    parser.add_argument('--companies', default=None, help='Comma-separated company names to filter (e.g., "ASUS,NVIDIA")')
    parser.add_argument('--no-confirm', action='store_true', help='Skip interactive confirmation prompt')
    args = parser.parse_args()

    _main_start = time.time()
    _ts = datetime.now().strftime('%Y%m%d_%H%M')
    _log_txt = f'logs/run_{_ts}.txt'
    _log_json = f'logs/run_{_ts}.json'
    _tee = TeeLogger(_log_txt)
    print(f"📋 Log 記錄至：{_log_txt}")

    # 載入配置
    config = load_config()
    scraper_config = getattr(config, 'SCRAPER_CONFIG', {}) if config else {}
    output_config = getattr(config, 'OUTPUT_CONFIG', {}) if config else {}
    mode = args.mode or scraper_config.get('mode', 'manual')
    task = args.task or 'matched'  # matched or baseline
    source_mode_filter = args.source_mode or 'all'
    company_filter = args.companies.split(',') if args.companies else None
    headless = scraper_config.get('headless', False)
    _base_output = output_config.get('filename', 'data/glassdoor_ratings.xlsx')
    _stem, _ext = os.path.splitext(_base_output)
    output_file = f"{_stem}_{_ts}{_ext}"

    # 掃描 data/ 目錄中的 URL 清單檔案（office/country/scan）
    # 根據 source_mode_filter 和 company_filter 過濾檔案
    matched_files = sorted(
        glob.glob('data/*_office.json') +
        glob.glob('data/*_country.json') +
        glob.glob('data/*_city.json') +
        glob.glob('data/*_scan.json')
    )

    # Apply source mode filter
    if source_mode_filter != 'all':
        filtered = []
        for f in matched_files:
            basename = os.path.basename(f)
            if source_mode_filter == 'country' and basename.endswith('_country.json'):
                filtered.append(f)
            elif source_mode_filter == 'office' and basename.endswith('_office.json'):
                filtered.append(f)
            elif source_mode_filter == 'city' and basename.endswith('_city.json'):
                filtered.append(f)
            elif source_mode_filter == 'scan' and basename.endswith('_scan.json'):
                filtered.append(f)
        matched_files = filtered
        print(f"🔍 Source Mode Filter: {source_mode_filter} -> {len(matched_files)} files selected")

    # Apply company name filter
    if company_filter:
        company_names = [c.strip().lower() for c in company_filter]
        filtered = []
        for f in matched_files:
            basename = os.path.basename(f).lower()
            # Match company name in filename (e.g., nvidia_office.json)
            for name in company_names:
                if name in basename or basename.startswith(name.replace(' ', '_')):
                    filtered.append(f)
                    break
        matched_files = filtered
        print(f"🔍 Company Filter: {company_filter} -> {len(matched_files)} files selected")

    if mode == 'manual' and not args.no_confirm:
        print("\n" + "="*60)
        print("手動模式 - 使用已登入的 Chrome")
        print("="*60)
        print("\n請確認：")
        print("1. 已執行 啟動Chrome.bat")
        print("2. 已在 Chrome 中登入 Glassdoor")
        print("\n如果還沒準備好，請按 Ctrl+C 取消")
        print("準備好後按 Enter 繼續...")
        input()

    include_baseline = (task == 'baseline') or (getattr(config, 'INCLUDE_BASELINE', False) if config else False)
    baseline_file = 'data/asus_office.json'
    wants_all_companies = company_filter is None
    wants_asus = wants_all_companies or any(c.strip().lower() == 'asus' for c in (company_filter or []))
    has_asus_file_selected = any(
        os.path.basename(f).lower() in ('asus_office.json', 'asus_country.json', 'asus_city.json')
        for f in matched_files
    )
    if task == 'matched' and wants_asus and not has_asus_file_selected:
        include_baseline = True
        if wants_all_companies:
            print("[INFO] All companies 模式自動加入 ASUS baseline")
        else:
            print("[INFO] 已選 ASUS，使用 baseline 補抓 ASUS 資料")
    # CLI --ports overrides config
    if args.ports:
        parallel_ports = [int(p.strip()) for p in args.ports.split(',')]
    else:
        parallel_ports = getattr(config, 'PARALLEL_PORTS', []) if config else []

    # 單一 port 或空清單 → 單一模式；多個 port → 平行模式
    use_parallel = len(parallel_ports) > 1

    try:
        if use_parallel:
            print(f"\n⚡ 平行模式：使用 {len(parallel_ports)} 個 Chrome（ports: {parallel_ports}）")
            print("請確認每個 port 的 Chrome 都已登入 Glassdoor\n")
            data = parallel_scrape(
                ports=parallel_ports,
                matched_files=matched_files,
                include_baseline=include_baseline,
                baseline_file=baseline_file,
                mode=mode,
                headless=headless,
                log_ts=_ts,
            )
        else:
            port = parallel_ports[0] if parallel_ports else 9222
            scraper = GlassdoorScraper(mode=mode, headless=headless, chrome_debugger_port=port)
            data = []

            if include_baseline and os.path.exists(baseline_file):
                asus_in_files = any(
                    os.path.basename(f) in ('asus_office.json', 'asus_country.json', 'asus_city.json')
                    for f in matched_files
                )
                if asus_in_files:
                    print("[INFO] asus_office.json/asus_country.json/asus_city.json 已在 URL list files 中，略過 baseline 避免重複")
                else:
                    print(f"\n[INCLUDE_BASELINE=True] 抓取基準公司 ASUS：{baseline_file}\n")
                    data += scraper.scrape_from_baseline_json(baseline_file, company_name='ASUS')

            if matched_files:
                print(f"\n找到 URL list JSON：{matched_files}")
                print(f"開始從 URL list JSON 抓取數據...\n")
                data += scraper.scrape_from_matched_json(matched_files)
            elif not include_baseline:
                if config and hasattr(config, 'COMPANY_URLS'):
                    companies = config.COMPANY_URLS
                    print(f"\n開始抓取 {len(companies)} 個公司的數據...\n")
                    data += scraper.scrape_multiple_companies(companies)
                else:
                    print("找不到 matched JSON 也沒有 config.py，請先執行 company_finder.py match")
                    return

            scraper.close()

        if data:
            # save_to_excel 不需要 driver，建一個不連 Chrome 的輕量實例
            saver = object.__new__(GlassdoorScraper)
            saver.mode = mode
            saver.save_to_excel(data, output_file)
            print(f"\n✓ 完成！共抓取 {len(data)} 筆數據")
        else:
            print("\n✗ 沒有成功抓取任何數據")

    except Exception as e:
        import traceback
        print(f"\n✗ 發生錯誤: {e}")
        traceback.print_exc()
        if mode == 'manual':
            print("\n請確認 Chrome 是否已用 --remote-debugging-port=9222 啟動")

    finally:
        elapsed = time.time() - _main_start
        print(f"\n⏱ 總執行時間：{elapsed//60:.0f}m {elapsed%60:.0f}s")
        try:
            save_run_report(data if 'data' in dir() else [], elapsed, _log_txt, _log_json)
        except Exception:
            pass
        _tee.restore()
        print(f"📋 Log 已儲存：{_log_txt}")


if __name__ == '__main__':
    main()
